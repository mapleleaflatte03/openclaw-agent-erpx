from __future__ import annotations

import csv
import mimetypes
import re
import shutil
import tempfile
import zipfile
from datetime import date, datetime, timedelta, timezone
from email import policy
from email.parser import BytesParser
from pathlib import Path
from typing import Any

import httpx
import pdfplumber
import pytesseract
from botocore.exceptions import BotoCoreError, ClientError
from celery import Task
from openpyxl import Workbook
from rapidfuzz import fuzz
from sqlalchemy import select
from sqlalchemy.exc import OperationalError

from openclaw_agent.agent_worker.celery_app import celery_app
from openclaw_agent.common.db import db_session, make_engine
from openclaw_agent.common.erpx_client import ErpXClient
from openclaw_agent.common.logging import configure_logging, get_logger
from openclaw_agent.common.models import (
    AgentAttachment,
    AgentAuditLog,
    AgentCloseTask,
    AgentContractCase,
    AgentEmailThread,
    AgentErpXLink,
    AgentEvidencePack,
    AgentException,
    AgentExport,
    AgentExtractedText,
    AgentKbDoc,
    AgentLog,
    AgentObligation,
    AgentObligationEvidence,
    AgentProposal,
    AgentReminderLog,
    AgentRun,
    AgentSourceFile,
    AgentTask,
)
from openclaw_agent.common.settings import get_settings
from openclaw_agent.common.storage import (
    download_file,
    parse_s3_uri,
    sha256_file,
    upload_file,
)
from openclaw_agent.common.utils import (
    json_dumps_canonical,
    make_idempotency_key,
    new_uuid,
    sha256_text,
    utcnow,
)

settings = get_settings()
configure_logging(settings.log_level)
log = get_logger("agent-worker")
engine = make_engine(settings.agent_db_dsn)


def _is_transient_error(e: Exception) -> bool:
    return isinstance(
        e,
        (
            httpx.TimeoutException,
            httpx.TransportError,
            BotoCoreError,
            ClientError,
            OperationalError,
        ),
    )


def _db_log(run_id: str, task_id: str | None, level: str, message: str, context: dict | None = None) -> None:
    with db_session(engine) as s:
        s.add(
            AgentLog(
                log_id=new_uuid(),
                run_id=run_id,
                task_id=task_id,
                level=level,
                message=message,
                context=context or None,
            )
        )


def _update_run(run_id: str, **fields: Any) -> None:
    with db_session(engine) as s:
        r = s.get(AgentRun, run_id)
        if not r:
            raise RuntimeError(f"run not found: {run_id}")
        for k, v in fields.items():
            setattr(r, k, v)


def _get_task_by_name(s, run_id: str, task_name: str) -> AgentTask | None:
    return s.execute(
        select(AgentTask).where((AgentTask.run_id == run_id) & (AgentTask.task_name == task_name))
    ).scalar_one_or_none()


def _task_start(run_id: str, task_name: str, input_ref: dict | None = None) -> str:
    with db_session(engine) as s:
        t = _get_task_by_name(s, run_id, task_name)
        if not t:
            t = AgentTask(
                task_id=new_uuid(),
                run_id=run_id,
                task_name=task_name,
                status="running",
                input_ref=input_ref,
                output_ref=None,
                error=None,
                started_at=utcnow(),
                finished_at=None,
            )
            s.add(t)
        else:
            t.status = "running"
            t.started_at = utcnow()
            t.error = None
        return t.task_id


def _task_finish(run_id: str, task_name: str, status: str, output_ref: dict | None = None, error: str | None = None) -> None:
    with db_session(engine) as s:
        t = _get_task_by_name(s, run_id, task_name)
        if not t:
            return
        t.status = status
        t.output_ref = output_ref
        t.error = error
        t.finished_at = utcnow()


def _safe_period_from_date_str(d: str | None) -> str | None:
    if not d:
        return None
    try:
        dt = datetime.fromisoformat(d.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m")
    except Exception:
        return None


def _extract_pdf_text(path: str) -> str:
    text = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            if t.strip():
                text.append(t)
    return "\n".join(text)


def _ocr_image(path: str) -> str:
    # pytesseract requires the tesseract binary installed in the container/host.
    return pytesseract.image_to_string(path, lang="eng+vie", timeout=settings.ocr_timeout_seconds)


def _ocr_pdf(path: str, max_pages: int) -> str:
    # Requires poppler utils (`pdftoppm`) inside the container.
    from pdf2image import convert_from_path

    pages = convert_from_path(path, first_page=1, last_page=max(max_pages, 1))
    texts: list[str] = []
    for img in pages:
        texts.append(pytesseract.image_to_string(img, lang="eng+vie", timeout=settings.ocr_timeout_seconds))
    return "\n".join(t for t in texts if t.strip())


def _parse_doc_keys(text: str) -> dict[str, Any]:
    # Minimal, deterministic rules for demo/golden tests.
    norm = " ".join(text.split())
    out: dict[str, Any] = {}

    m = re.search(r"(?:S[oố]\s*h[oó]a\s*đ[oơ]n|Invoice\s*No)\s*[:#]?\s*([A-Z0-9/-]+)", norm, re.I)
    if m:
        out["invoice_no"] = m.group(1).strip()

    m = re.search(r"(?:MST|Tax\s*ID)\s*[:#]?\s*([0-9]{10,13})", norm, re.I)
    if m:
        out["tax_id"] = m.group(1).strip()

    m = re.search(r"(?:Ng[aà]y|Date)\s*[:#]?\s*([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{4})", norm, re.I)
    if m:
        out["doc_date"] = m.group(1).strip()

    m = re.search(r"(?:T[oổng]\s*ti[eề]n|Total)\s*[:#]?\s*([0-9][0-9\\.,]*)", norm, re.I)
    if m:
        out["amount_raw"] = m.group(1).strip()

    m = re.search(r"(?:M[aã]\s*KH|Customer\s*Code)\s*[:#]?\s*([A-Z0-9\\-]+)", norm, re.I)
    if m:
        out["customer_code"] = m.group(1).strip()

    return out


def _match_invoice(parsed: dict[str, Any], invoices: list[dict], threshold: float) -> tuple[dict | None, float]:
    invoice_no = parsed.get("invoice_no")
    tax_id = parsed.get("tax_id")
    if invoice_no:
        for inv in invoices:
            if str(inv.get("invoice_no", "")).strip().upper() == str(invoice_no).strip().upper():
                if tax_id and str(inv.get("tax_id", "")).strip() != str(tax_id).strip():
                    # hard mismatch
                    continue
                return inv, 1.0

    best: dict | None = None
    best_score = 0.0
    # Fuzzy fallback: invoice_no similarity
    if invoice_no:
        for inv in invoices:
            score = fuzz.ratio(str(inv.get("invoice_no", "")), str(invoice_no)) / 100.0
            if tax_id and inv.get("tax_id") == tax_id:
                score += 0.05
            if score > best_score:
                best_score = score
                best = inv

    if best_score >= threshold:
        return best, best_score
    return None, best_score


def _ensure_export_unique(s, export_type: str, period: str, force_new_version: bool) -> int:
    rows = s.execute(
        select(AgentExport)
        .where((AgentExport.export_type == export_type) & (AgentExport.period == period))
        .order_by(AgentExport.version.desc())
    ).scalars().all()
    if not rows:
        return 1
    if force_new_version:
        return int(rows[0].version) + 1
    # idempotent: keep latest
    return int(rows[0].version)


def _csv_write(path: str, rows: list[dict[str, Any]]) -> None:
    if not rows:
        Path(path).write_text("", encoding="utf-8")
        return
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def _xlsx_vat_list(path: str, invoices: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "VAT_List"
    ws.append(["invoice_id", "invoice_no", "tax_id", "date", "amount", "customer_id", "status"])
    for inv in invoices:
        ws.append(
            [
                inv.get("invoice_id"),
                inv.get("invoice_no"),
                inv.get("tax_id"),
                inv.get("date"),
                inv.get("amount"),
                inv.get("customer_id"),
                inv.get("status"),
            ]
        )
    wb.save(path)


def _xlsx_working_papers(path: str, balances: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["period", balances.get("period")])
    ws.append(["gl_total", balances.get("gl_total")])
    ws.append(["ar_total", balances.get("ar_total")])
    ws.append(["ap_total", balances.get("ap_total")])

    ws2 = wb.create_sheet("AR_Aging")
    ws2.append(["customer_id", "invoice_id", "overdue_days", "amount"])
    for row in balances.get("ar_aging", []):
        ws2.append([row.get("customer_id"), row.get("invoice_id"), row.get("overdue_days"), row.get("amount")])

    wb.save(path)


def _write_summary_md(path: str, title: str, lines: list[str]) -> None:
    Path(path).write_text("# " + title + "\n\n" + "\n".join(lines) + "\n", encoding="utf-8")


def _send_email(to_addr: str, subject: str, body: str) -> None:
    # Optional; if SMTP not configured, caller should skip.
    import smtplib
    from email.message import EmailMessage

    msg = EmailMessage()
    msg["From"] = settings.smtp_from
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg.set_content(body)

    if settings.smtp_tls:
        server = smtplib.SMTP(settings.smtp_host, settings.smtp_port, timeout=10)
        server.starttls()
    else:
        server = smtplib.SMTP(settings.smtp_host, settings.smtp_port, timeout=10)

    try:
        if settings.smtp_user:
            server.login(settings.smtp_user, settings.smtp_password)
        server.send_message(msg)
    finally:
        server.quit()


@celery_app.task(name="openclaw_agent.agent_worker.tasks.dispatch_run", bind=True)
def dispatch_run(self: Task, run_id: str) -> dict[str, Any]:
    with db_session(engine) as s:
        run = s.get(AgentRun, run_id)
        if not run:
            raise RuntimeError(f"run not found: {run_id}")
        if run.status in {"success", "failed", "canceled"}:
            return {"run_id": run_id, "status": run.status}
        run_type = run.run_type
        run.status = "running"
        run.started_at = utcnow()

    _db_log(run_id, None, "info", "run_started", {"run_id": run_id})

    try:
        if run_type == "attachment":
            stats = _wf_attachment(run_id)
        elif run_type == "tax_export":
            stats = _wf_tax_export(run_id)
        elif run_type == "working_papers":
            stats = _wf_working_papers(run_id)
        elif run_type == "soft_checks":
            stats = _wf_soft_checks(run_id)
        elif run_type == "ar_dunning":
            stats = _wf_ar_dunning(run_id)
        elif run_type == "close_checklist":
            stats = _wf_close_checklist(run_id)
        elif run_type == "evidence_pack":
            stats = _wf_evidence_pack(run_id)
        elif run_type == "kb_index":
            stats = _wf_kb_index(run_id)
        elif run_type == "contract_obligation":
            stats = _wf_contract_obligation(run_id)
        else:
            raise RuntimeError(f"unsupported run_type: {run_type}")

        _update_run(run_id, status="success", finished_at=utcnow(), stats=stats)
        _db_log(run_id, None, "info", "run_success", {"stats": stats})
        return {"run_id": run_id, "status": "success", "stats": stats}
    except Exception as e:
        max_attempts = max(1, int(settings.task_retry_max_attempts))
        max_retries = max_attempts - 1
        retries = int(getattr(getattr(self, "request", None), "retries", 0) or 0)
        if _is_transient_error(e) and retries < max_retries:
            countdown = int(settings.task_retry_backoff_seconds) * (2 ** retries)
            _db_log(
                run_id,
                None,
                "warn",
                "run_retrying",
                {
                    "error": str(e),
                    "retries": retries,
                    "max_retries": max_retries,
                    "retry_in_seconds": countdown,
                },
            )
            raise self.retry(exc=e, countdown=countdown, max_retries=max_retries) from e

        _update_run(run_id, status="failed", finished_at=utcnow(), stats={"error": str(e)})
        _db_log(run_id, None, "error", "run_failed", {"error": str(e)})
        raise


def _run_payload(run_id: str) -> dict[str, Any]:
    with db_session(engine) as s:
        r = s.get(AgentRun, run_id)
        if not r:
            raise RuntimeError(f"run not found: {run_id}")
        return r.cursor_in or {}


def _wf_attachment(run_id: str) -> dict[str, Any]:
    payload = _run_payload(run_id)
    file_uri = payload.get("file_uri")
    if not file_uri:
        raise RuntimeError("payload.file_uri is required")

    workdir = Path(tempfile.mkdtemp(prefix=f"agent-{run_id}-"))
    try:
        t_id = _task_start(run_id, "extract_text", {"file_uri": file_uri})
        if str(file_uri).startswith("s3://"):
            ref = parse_s3_uri(file_uri)
            suffix = Path(ref.key).suffix or ".bin"
            local_path = str(workdir / f"input{suffix}")
            download_file(settings, ref, local_path)
        else:
            suffix = Path(str(file_uri)).suffix or ".bin"
            local_path = str(workdir / f"input{suffix}")
            shutil.copyfile(file_uri, local_path)

        file_hash = sha256_file(local_path)
        ext = Path(local_path).suffix.lower()
        if ext == ".pdf":
            text = _extract_pdf_text(local_path)
            if not text.strip():
                text = _ocr_pdf(local_path, max_pages=settings.ocr_pdf_max_pages)
        elif ext in {".png", ".jpg", ".jpeg", ".tif", ".tiff"}:
            text = _ocr_image(local_path)
        else:
            text = ""
        _task_finish(run_id, "extract_text", "success", {"file_hash": file_hash, "text_len": len(text)})
        _db_log(run_id, t_id, "info", "text_extracted", {"file_hash": file_hash, "text_len": len(text)})

        t_id = _task_start(run_id, "parse_keys", {"text_len": len(text)})
        parsed = _parse_doc_keys(text)
        _task_finish(run_id, "parse_keys", "success", {"parsed": parsed})
        _db_log(run_id, t_id, "info", "keys_parsed", {"parsed": parsed})

        t_id = _task_start(run_id, "match", parsed)
        client = ErpXClient(settings)
        try:
            period = payload.get("period") or _safe_period_from_date_str(payload.get("doc_ts"))
            # If document date is in dd/mm/yyyy -> derive period.
            if not period and parsed.get("doc_date"):
                try:
                    dd, mm, yyyy = re.split(r"[/-]", parsed["doc_date"])
                    period = f"{int(yyyy):04d}-{int(mm):02d}"
                except Exception:
                    period = None
            if not period:
                period = datetime.now(timezone.utc).strftime("%Y-%m")
            invoices = client.get_invoices(period=period)
        finally:
            client.close()

        inv, score = _match_invoice(parsed, invoices, settings.match_confidence_threshold)
        if not inv:
            # Safety-first: do not attach. Create exception.
            signature = sha256_text(json_dumps_canonical(["attachment_mismatch", file_hash, parsed]))[:64]
            with db_session(engine) as s:
                existing = s.execute(select(AgentException).where(AgentException.signature == signature)).scalar_one_or_none()
                if not existing:
                    s.add(
                        AgentException(
                            id=new_uuid(),
                            exception_type="attachment_mismatch",
                            severity="med",
                            erp_refs={"file_uri": file_uri},
                            summary="Could not confidently match attachment to ERP object",
                            details={"parsed": parsed, "confidence": score},
                            signature=signature,
                            run_id=run_id,
                        )
                    )
            _task_finish(run_id, "match", "failed", {"confidence": score}, "no confident match")
            raise RuntimeError(f"no confident match (score={score:.2f})")

        match_info = {"erp_object_type": "invoice", "erp_object_id": inv["invoice_id"], "confidence": score}
        _task_finish(run_id, "match", "success", match_info)
        _db_log(run_id, t_id, "info", "matched", match_info)

        t_id = _task_start(run_id, "attach", match_info)
        # Upload original file into attachments bucket (content-addressed)
        ext = ".pdf" if local_path.lower().endswith(".pdf") else Path(local_path).suffix or ".bin"
        key = f"{match_info['erp_object_type']}/{match_info['erp_object_id']}/{file_hash}{ext}"
        obj = upload_file(settings, settings.minio_bucket_attachments, key, local_path)

        with db_session(engine) as s:
            existing = s.execute(
                select(AgentAttachment).where(
                    (AgentAttachment.file_hash == file_hash)
                    & (AgentAttachment.erp_object_type == match_info["erp_object_type"])
                    & (AgentAttachment.erp_object_id == match_info["erp_object_id"])
                )
            ).scalar_one_or_none()
            if not existing:
                s.add(
                    AgentAttachment(
                        id=new_uuid(),
                        erp_object_type=match_info["erp_object_type"],
                        erp_object_id=match_info["erp_object_id"],
                        file_uri=obj.uri(),
                        file_hash=file_hash,
                        matched_by="rule" if score >= 0.99 else "ocr",
                        run_id=run_id,
                    )
                )

        _task_finish(run_id, "attach", "success", {"file_uri": obj.uri()})
        _db_log(run_id, t_id, "info", "attached", {"file_uri": obj.uri()})

        _update_run(run_id, cursor_out={"file_hash": file_hash, **match_info})
        return {"attachments": 1, "matched_confidence": score}
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def _wf_tax_export(run_id: str) -> dict[str, Any]:
    payload = _run_payload(run_id)
    period = payload.get("period")
    if not period:
        raise RuntimeError("payload.period is required (YYYY-MM)")
    force_new = bool(payload.get("force_new_version", False))

    _task_start(run_id, "pull_invoices", {"period": period})
    client = ErpXClient(settings)
    try:
        invoices = client.get_invoices(period=period)
    finally:
        client.close()
    _task_finish(run_id, "pull_invoices", "success", {"count": len(invoices)})

    _task_start(run_id, "validate", {"count": len(invoices)})
    errors = []
    required = ["invoice_id", "invoice_no", "tax_id", "date", "amount"]
    for inv in invoices:
        missing = [k for k in required if not inv.get(k)]
        if missing:
            errors.append({"invoice_id": inv.get("invoice_id"), "missing": missing})
    if errors:
        signature = sha256_text(json_dumps_canonical(["vat_export_missing_fields", period, errors]))[:64]
        with db_session(engine) as s:
            if not s.execute(select(AgentException).where(AgentException.signature == signature)).scalar_one_or_none():
                s.add(
                    AgentException(
                        id=new_uuid(),
                        exception_type="vat_export_missing_fields",
                        severity="high",
                        erp_refs={"period": period},
                        summary="Invoices missing required fields for VAT export",
                        details={"errors": errors[:50]},
                        signature=signature,
                        run_id=run_id,
                    )
                )
        _task_finish(run_id, "validate", "failed", {"errors": len(errors)}, "missing fields")
        raise RuntimeError(f"VAT export validation failed: {len(errors)} invoices missing fields")
    _task_finish(run_id, "validate", "success", {"validated": len(invoices)})

    _task_start(run_id, "export_xlsx", {"period": period})
    workdir = Path(tempfile.mkdtemp(prefix=f"agent-{run_id}-"))
    try:
        with db_session(engine) as s:
            version = _ensure_export_unique(s, "vat_list", period, force_new)
            existing = s.execute(
                select(AgentExport).where(
                    (AgentExport.export_type == "vat_list")
                    & (AgentExport.period == period)
                    & (AgentExport.version == version)
                )
            ).scalar_one_or_none()
            if existing and not force_new:
                _task_finish(run_id, "export_xlsx", "success", {"file_uri": existing.file_uri, "version": version})
                _update_run(run_id, cursor_out={"period": period, "file_uri": existing.file_uri, "version": version})
                return {"export": 0, "reused": 1, "version": version}

        out_path = str(workdir / f"vat_list_{period}_v{version}.xlsx")
        _xlsx_vat_list(out_path, invoices)
        checksum = sha256_file(out_path)
        key = f"vat_list/{period}/vat_list_v{version}.xlsx"
        obj = upload_file(settings, settings.minio_bucket_exports, key, out_path)

        with db_session(engine) as s:
            s.add(
                AgentExport(
                    id=new_uuid(),
                    export_type="vat_list",
                    period=period,
                    version=version,
                    file_uri=obj.uri(),
                    checksum=checksum,
                    run_id=run_id,
                )
            )

        _task_finish(run_id, "export_xlsx", "success", {"file_uri": obj.uri(), "version": version})
        _update_run(run_id, cursor_out={"period": period, "file_uri": obj.uri(), "version": version})
        return {"export": 1, "version": version}
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def _wf_working_papers(run_id: str) -> dict[str, Any]:
    payload = _run_payload(run_id)
    period = payload.get("period")
    if not period:
        raise RuntimeError("payload.period is required (YYYY-MM)")
    force_new = bool(payload.get("force_new_version", False))

    _task_start(run_id, "pull_balances", {"period": period})
    client = ErpXClient(settings)
    try:
        # MVP: use AR aging as "balances" payload for template.
        ar = client.get_ar_aging(as_of=f"{period}-28")
    finally:
        client.close()
    balances = {
        "period": period,
        "gl_total": 0,
        "ar_total": sum(float(x.get("amount", 0) or 0) for x in ar),
        "ap_total": 0,
        "ar_aging": ar,
    }
    _task_finish(run_id, "pull_balances", "success", {"ar_rows": len(ar)})

    _task_start(run_id, "fill_templates", {"period": period})
    _task_finish(run_id, "fill_templates", "success")

    _task_start(run_id, "export_bundle", {"period": period})
    workdir = Path(tempfile.mkdtemp(prefix=f"agent-{run_id}-"))
    try:
        with db_session(engine) as s:
            version = _ensure_export_unique(s, "working_paper", period, force_new)
            existing = s.execute(
                select(AgentExport).where(
                    (AgentExport.export_type == "working_paper")
                    & (AgentExport.period == period)
                    & (AgentExport.version == version)
                )
            ).scalar_one_or_none()
            if existing and not force_new:
                _task_finish(run_id, "export_bundle", "success", {"file_uri": existing.file_uri, "version": version})
                _update_run(run_id, cursor_out={"period": period, "file_uri": existing.file_uri, "version": version})
                return {"export": 0, "reused": 1, "version": version}

        out_path = str(workdir / f"working_papers_{period}_v{version}.xlsx")
        _xlsx_working_papers(out_path, balances)
        checksum = sha256_file(out_path)
        key = f"working_paper/{period}/working_papers_v{version}.xlsx"
        obj = upload_file(settings, settings.minio_bucket_exports, key, out_path)

        with db_session(engine) as s:
            s.add(
                AgentExport(
                    id=new_uuid(),
                    export_type="working_paper",
                    period=period,
                    version=version,
                    file_uri=obj.uri(),
                    checksum=checksum,
                    run_id=run_id,
                )
            )

        _task_finish(run_id, "export_bundle", "success", {"file_uri": obj.uri(), "version": version})
        _update_run(run_id, cursor_out={"period": period, "file_uri": obj.uri(), "version": version})
        return {"export": 1, "version": version}
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def _wf_soft_checks(run_id: str) -> dict[str, Any]:
    payload = _run_payload(run_id)
    updated_after = payload.get("updated_after")  # ISO string
    period = payload.get("period") or datetime.now(timezone.utc).strftime("%Y-%m")
    force_new = bool(payload.get("force_new_version", False))

    _task_start(run_id, "pull_delta", {"updated_after": updated_after})
    client = ErpXClient(settings)
    try:
        vouchers = client.get_vouchers(updated_after=updated_after)
        journals = client.get_journals(updated_after=updated_after)
        invoices = client.get_invoices(period=period)
    finally:
        client.close()
    _task_finish(
        run_id, "pull_delta", "success", {"vouchers": len(vouchers), "journals": len(journals), "invoices": len(invoices)}
    )

    _task_start(run_id, "checks", {"period": period})
    exceptions: list[AgentException] = []

    # Check: voucher missing attachment flag
    for v in vouchers:
        if not v.get("has_attachment", True):
            signature = sha256_text(json_dumps_canonical(["missing_attachment", v.get("voucher_id"), period]))[:64]
            exceptions.append(
                AgentException(
                    id=new_uuid(),
                    exception_type="missing_attachment",
                    severity="med",
                    erp_refs={"voucher_id": v.get("voucher_id")},
                    summary="Voucher missing supporting attachment",
                    details={"voucher": v},
                    signature=signature,
                    run_id=run_id,
                )
            )

    # Check: journal out of balance
    for j in journals:
        if float(j.get("debit_total", 0) or 0) != float(j.get("credit_total", 0) or 0):
            signature = sha256_text(json_dumps_canonical(["journal_imbalanced", j.get("journal_id"), period]))[:64]
            exceptions.append(
                AgentException(
                    id=new_uuid(),
                    exception_type="journal_imbalanced",
                    severity="high",
                    erp_refs={"journal_id": j.get("journal_id")},
                    summary="Journal entry is not balanced (debit != credit)",
                    details={"journal": j},
                    signature=signature,
                    run_id=run_id,
                )
            )

    # Check: overdue invoices
    today = date.today()
    for inv in invoices:
        if inv.get("status") == "unpaid" and inv.get("due_date"):
            try:
                due = date.fromisoformat(inv["due_date"])
                if due < today:
                    signature = sha256_text(json_dumps_canonical(["invoice_overdue", inv.get("invoice_id"), period]))[:64]
                    exceptions.append(
                        AgentException(
                            id=new_uuid(),
                            exception_type="invoice_overdue",
                            severity="low",
                            erp_refs={"invoice_id": inv.get("invoice_id")},
                            summary="Invoice is overdue",
                            details={"invoice": inv, "overdue_days": (today - due).days},
                            signature=signature,
                            run_id=run_id,
                        )
                    )
            except Exception:
                continue

    with db_session(engine) as s:
        inserted = 0
        for ex in exceptions:
            if not s.execute(select(AgentException).where(AgentException.signature == ex.signature)).scalar_one_or_none():
                s.add(ex)
                inserted += 1

    _task_finish(run_id, "checks", "success", {"exceptions": len(exceptions)})

    # Export report
    _task_start(run_id, "export_report", {"period": period, "exceptions": len(exceptions)})
    workdir = Path(tempfile.mkdtemp(prefix=f"agent-{run_id}-"))
    try:
        with db_session(engine) as s:
            version = _ensure_export_unique(s, "soft_checks", period, force_new)
            existing = s.execute(
                select(AgentExport).where(
                    (AgentExport.export_type == "soft_checks")
                    & (AgentExport.period == period)
                    & (AgentExport.version == version)
                )
            ).scalar_one_or_none()
            if existing and not force_new:
                _task_finish(run_id, "export_report", "success", {"file_uri": existing.file_uri, "version": version})
                _update_run(
                    run_id,
                    cursor_out={"period": period, "exceptions": len(exceptions), "report_uri": existing.file_uri},
                )
                return {"exceptions": len(exceptions), "reused_report": 1}

        report_path = str(workdir / f"soft_checks_{period}.csv")
        _csv_write(
            report_path,
            [
                {
                    "exception_type": e.exception_type,
                    "severity": e.severity,
                    "summary": e.summary,
                    "signature": e.signature,
                }
                for e in exceptions
            ],
        )
        checksum = sha256_file(report_path)
        key = f"soft_checks/{period}/soft_checks_v{version}.csv"
        obj = upload_file(settings, settings.minio_bucket_exports, key, report_path)

        with db_session(engine) as s:
            s.add(
                AgentExport(
                    id=new_uuid(),
                    export_type="soft_checks",
                    period=period,
                    version=version,
                    file_uri=obj.uri(),
                    checksum=checksum,
                    run_id=run_id,
                )
            )

        _task_finish(run_id, "export_report", "success", {"file_uri": obj.uri(), "version": version})
        _update_run(run_id, cursor_out={"period": period, "exceptions": len(exceptions), "report_uri": obj.uri()})
        return {"exceptions": len(exceptions)}
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def _wf_ar_dunning(run_id: str) -> dict[str, Any]:
    payload = _run_payload(run_id)
    as_of = payload.get("as_of") or date.today().isoformat()
    window_days = int(payload.get("policy_window_days", 30))

    _task_start(run_id, "pull_ar_aging", {"as_of": as_of})
    client = ErpXClient(settings)
    try:
        rows = client.get_ar_aging(as_of=as_of)
    finally:
        client.close()
    _task_finish(run_id, "pull_ar_aging", "success", {"rows": len(rows)})

    _task_start(run_id, "apply_policy", {"rows": len(rows)})
    candidates = []
    for r in rows:
        overdue = int(r.get("overdue_days") or 0)
        stage = 0
        if overdue >= 7:
            stage = 1
        if overdue >= 14:
            stage = 2
        if overdue >= 30:
            stage = 3
        if stage:
            candidates.append({**r, "stage": stage})
    _task_finish(run_id, "apply_policy", "success", {"candidates": len(candidates)})

    _task_start(run_id, "notify", {"candidates": len(candidates)})
    sent = 0
    skipped = 0
    now = utcnow()
    cutoff = now - timedelta(days=window_days)
    with db_session(engine) as s:
        for c in candidates:
            invoice_id = str(c.get("invoice_id") or "")
            if not invoice_id:
                continue
            stage = int(c["stage"])
            sent_to = str(c.get("email") or c.get("customer_email") or "internal")
            # Idempotency: only 1 reminder per invoice_id+stage within policy window.
            existing_recent = (
                s.execute(
                    select(AgentReminderLog)
                    .where(
                        (AgentReminderLog.invoice_id == invoice_id)
                        & (AgentReminderLog.reminder_stage == stage)
                        & (AgentReminderLog.sent_at >= cutoff)
                    )
                    .order_by(AgentReminderLog.sent_at.desc())
                    .limit(1)
                )
                .scalar_one_or_none()
            )
            if existing_recent:
                skipped += 1
                continue

            # Stable key for uniqueness in DB (handles concurrent sends).
            window_bucket = int(now.date().toordinal() // max(window_days, 1))
            policy_key = make_idempotency_key("ar_dunning", invoice_id, stage, window_days, window_bucket)

            existing = s.execute(select(AgentReminderLog).where(AgentReminderLog.policy_key == policy_key)).scalar_one_or_none()
            if existing:
                skipped += 1
                continue

            # send (email optional)
            if settings.smtp_host and sent_to != "internal":
                _send_email(
                    sent_to,
                    subject=f"[AR Reminder] Invoice {invoice_id} - Stage {stage}",
                    body=f"Reminder stage {stage} for invoice {invoice_id}. Overdue days: {c.get('overdue_days')}",
                )
                channel = "email"
            else:
                channel = "internal"

            s.add(
                AgentReminderLog(
                    id=new_uuid(),
                    customer_id=str(c.get("customer_id") or ""),
                    invoice_id=invoice_id,
                    reminder_stage=stage,
                    channel=channel,
                    sent_to=sent_to,
                    sent_at=now,
                    run_id=run_id,
                    policy_key=policy_key,
                )
            )
            sent += 1

    _task_finish(run_id, "notify", "success", {"sent": sent, "skipped": skipped})
    _update_run(run_id, cursor_out={"as_of": as_of, "sent": sent, "skipped": skipped})
    return {"sent": sent, "skipped": skipped}


def _wf_close_checklist(run_id: str) -> dict[str, Any]:
    payload = _run_payload(run_id)
    period = payload.get("period")
    if not period:
        raise RuntimeError("payload.period is required (YYYY-MM)")

    _task_start(run_id, "pull_close_calendar", {"period": period})
    client = ErpXClient(settings)
    try:
        items = client.get_close_calendar(period=period)
    finally:
        client.close()
    _task_finish(run_id, "pull_close_calendar", "success", {"items": len(items)})

    _task_start(run_id, "upsert_close_tasks", {"items": len(items)})
    upserted = 0
    with db_session(engine) as s:
        for it in items:
            task_name = str(it.get("task_name"))
            due_date_str = it.get("due_date")
            if not task_name or not due_date_str:
                continue
            due = date.fromisoformat(due_date_str)
            existing = s.execute(
                select(AgentCloseTask).where(
                    (AgentCloseTask.period == period) & (AgentCloseTask.task_name == task_name)
                )
            ).scalar_one_or_none()
            if existing:
                existing.owner_user_id = it.get("owner_user_id")
                existing.due_date = due
                existing.status = existing.status or "todo"
            else:
                s.add(
                    AgentCloseTask(
                        id=new_uuid(),
                        period=period,
                        task_name=task_name,
                        owner_user_id=it.get("owner_user_id"),
                        due_date=due,
                        status="todo",
                        last_nudged_at=None,
                    )
                )
            upserted += 1
    _task_finish(run_id, "upsert_close_tasks", "success", {"upserted": upserted})

    _task_start(run_id, "nudge", {"period": period})
    nudged = 0
    now = utcnow()
    with db_session(engine) as s:
        tasks = s.execute(select(AgentCloseTask).where(AgentCloseTask.period == period)).scalars().all()
        for t in tasks:
            if t.status in {"done"}:
                continue
            if t.due_date <= date.today() + timedelta(days=2):
                # Nudge at most once per day
                if t.last_nudged_at and t.last_nudged_at.date() == date.today():
                    continue
                t.last_nudged_at = now
                nudged += 1
    _task_finish(run_id, "nudge", "success", {"nudged": nudged})
    _update_run(run_id, cursor_out={"period": period, "upserted": upserted, "nudged": nudged})
    return {"upserted": upserted, "nudged": nudged}


def _wf_evidence_pack(run_id: str) -> dict[str, Any]:
    payload = _run_payload(run_id)
    exception_id = payload.get("exception_id")
    issue_id = payload.get("issue_id")
    if not exception_id and not issue_id:
        raise RuntimeError("payload.exception_id or payload.issue_id is required")
    issue_key = f"exception:{exception_id}" if exception_id else f"issue:{issue_id}"

    _task_start(run_id, "collect", {"issue_key": issue_key})
    with db_session(engine) as s:
        existing = s.execute(
            select(AgentEvidencePack).where((AgentEvidencePack.issue_key == issue_key) & (AgentEvidencePack.version == 1))
        ).scalar_one_or_none()
        if existing:
            _task_finish(run_id, "collect", "success", {"reused": True})
            _update_run(run_id, cursor_out={"issue_key": issue_key, "pack_uri": existing.pack_uri})
            return {"reused": 1}

        ex = s.get(AgentException, exception_id) if exception_id else None
        refs = (ex.erp_refs if ex else {}) if exception_id else {"issue_id": issue_id}

    _task_finish(run_id, "collect", "success", {"refs": refs})

    _task_start(run_id, "pack", {"issue_key": issue_key})
    workdir = Path(tempfile.mkdtemp(prefix=f"agent-{run_id}-"))
    try:
        summary_path = str(workdir / "summary.md")
        _write_summary_md(summary_path, "Evidence Pack", [f"- issue_key: {issue_key}", f"- refs: {refs}"])

        zip_path = str(workdir / "evidence.zip")
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
            z.write(summary_path, arcname="summary.md")

        checksum = sha256_file(zip_path)
        key = f"evidence/{issue_key}/v1/evidence.zip"
        obj = upload_file(settings, settings.minio_bucket_evidence, key, zip_path)

        with db_session(engine) as s:
            s.add(
                AgentEvidencePack(
                    id=new_uuid(),
                    issue_key=issue_key,
                    version=1,
                    pack_uri=obj.uri(),
                    index_json={"checksum": checksum, "refs": refs},
                    run_id=run_id,
                )
            )

        _task_finish(run_id, "pack", "success", {"pack_uri": obj.uri()})
        _task_start(run_id, "register", {"pack_uri": obj.uri()})
        _task_finish(run_id, "register", "success")

        _update_run(run_id, cursor_out={"issue_key": issue_key, "pack_uri": obj.uri()})
        return {"packed": 1}
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def _wf_kb_index(run_id: str) -> dict[str, Any]:
    payload = _run_payload(run_id)
    file_uri = payload.get("file_uri")
    if not file_uri:
        raise RuntimeError("payload.file_uri is required")

    workdir = Path(tempfile.mkdtemp(prefix=f"agent-{run_id}-"))
    try:
        _task_start(run_id, "extract_text", {"file_uri": file_uri})
        if str(file_uri).startswith("s3://"):
            ref = parse_s3_uri(file_uri)
            suffix = Path(ref.key).suffix or ".bin"
            local_path = str(workdir / f"kb_input{suffix}")
            download_file(settings, ref, local_path)
        else:
            suffix = Path(str(file_uri)).suffix or ".bin"
            local_path = str(workdir / f"kb_input{suffix}")
            shutil.copyfile(file_uri, local_path)

        file_hash = sha256_file(local_path)
        ext = Path(local_path).suffix.lower()
        if ext == ".pdf":
            text = _extract_pdf_text(local_path)
            if not text.strip():
                text = _ocr_pdf(local_path, max_pages=settings.ocr_pdf_max_pages)
        elif ext in {".png", ".jpg", ".jpeg", ".tif", ".tiff"}:
            text = _ocr_image(local_path)
        else:
            text = ""
        _task_finish(run_id, "extract_text", "success", {"file_hash": file_hash, "text_len": len(text)})

        _task_start(run_id, "extract_meta", {"text_len": len(text)})
        title = payload.get("title") or (text.strip().splitlines()[0][:120] if text.strip() else "Untitled")
        doc_type = payload.get("doc_type") or "process"
        version = payload.get("version") or "v1"
        effective_date = payload.get("effective_date")
        meta = {"keywords": list({w.lower() for w in re.findall(r"[A-Za-z0-9]{4,}", text)[:50]})}
        _task_finish(run_id, "extract_meta", "success", {"title": title, "doc_type": doc_type, "version": version})

        _task_start(run_id, "index", {"keywords": len(meta["keywords"])})
        _task_finish(run_id, "index", "success")

        _task_start(run_id, "register", {"file_hash": file_hash})
        # store extracted text
        text_path = str(workdir / "doc.txt")
        Path(text_path).write_text(text, encoding="utf-8")
        key = f"kb/text/{file_hash}.txt"
        obj = upload_file(settings, settings.minio_bucket_kb, key, text_path, content_type="text/plain")

        with db_session(engine) as s:
            existing = s.execute(
                select(AgentKbDoc).where((AgentKbDoc.file_hash == file_hash) & (AgentKbDoc.version == version))
            ).scalar_one_or_none()
            if not existing:
                s.add(
                    AgentKbDoc(
                        id=new_uuid(),
                        doc_type=doc_type,
                        title=title,
                        version=version,
                        effective_date=date.fromisoformat(effective_date) if effective_date else None,
                        source_uri=file_uri,
                        text_uri=obj.uri(),
                        indexed_at=utcnow(),
                        file_hash=file_hash,
                        meta=meta,
                    )
                )

        _task_finish(run_id, "register", "success", {"text_uri": obj.uri()})
        _update_run(run_id, cursor_out={"file_hash": file_hash, "text_uri": obj.uri()})
        return {"indexed": 1}
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def _clean_email_text(text: str) -> str:
    # Keep deterministic and conservative: remove quoted replies and common separators.
    out_lines: list[str] = []
    for line in text.splitlines():
        s = line.rstrip()
        if s.strip().startswith(">"):
            continue
        if s.strip().lower().startswith("-----original message-----"):
            break
        out_lines.append(s)
    return "\n".join(out_lines).strip()


def _parse_email_file(path: str) -> tuple[str | None, str | None, list[str] | None, str]:
    ext = Path(path).suffix.lower()
    if ext == ".eml":
        msg = BytesParser(policy=policy.default).parsebytes(Path(path).read_bytes())
        subject = str(msg.get("subject") or "") or None
        from_addr = str(msg.get("from") or "") or None
        to_addrs = [str(v) for v in (msg.get_all("to") or [])] or None

        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    try:
                        body = part.get_content() or ""
                    except Exception:
                        body = ""
                    break
        else:
            try:
                body = msg.get_content() or ""
            except Exception:
                body = ""

        return subject, from_addr, to_addrs, _clean_email_text(body)

    raw = Path(path).read_text(encoding="utf-8", errors="ignore")
    return None, None, None, _clean_email_text(raw)


def _try_parse_date_any(s: str) -> date | None:
    s = s.strip()
    try:
        return date.fromisoformat(s)
    except Exception:
        pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def _extract_contract_meta(text: str) -> dict[str, str]:
    norm = " ".join(text.split())
    out: dict[str, str] = {}

    m = re.search(r"(?:MST|Tax\s*ID)\s*[:#]?\s*([0-9]{10,13})", norm, re.I)
    if m:
        out["partner_tax_id"] = m.group(1).strip()

    m = re.search(
        r"(?:H[oơ]p\s*đ[oồ]ng\s*s[oố]|S[oố]\s*HĐ|Contract\s*(?:No|Code))\s*[:#]?\s*([A-Z0-9/-]+)",
        norm,
        re.I,
    )
    if m:
        out["contract_code"] = m.group(1).strip()

    return out


def _parse_percent(s: str) -> float | None:
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", s)
    if not m:
        return None
    try:
        return float(m.group(1))
    except Exception:
        return None


def _extract_obligation_candidates(text: str) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for raw_line in text.splitlines():
        line = " ".join(raw_line.strip().split())
        if not line:
            continue
        lower = line.lower()
        is_discount = "discount" in lower or "chiết khấu" in lower or "chiet khau" in lower
        is_penalty = "penalty" in lower or "phạt" in lower or "phat" in lower

        currency = None
        m_cur = re.search(r"\b(VND|USD|EUR)\b", line)
        if m_cur:
            currency = m_cur.group(1).upper()

        m_due = re.search(r"(\d{4}-\d{2}-\d{2}|\d{1,2}[/-]\d{1,2}[/-]\d{4})", line)
        due_date = _try_parse_date_any(m_due.group(1)) if m_due else None
        m_within = re.search(r"(?:within|trong\s*v[oò]ng)\s*(\d{1,3})\s*(?:days|ng[aà]y)", line, re.I)
        within_days = int(m_within.group(1)) if m_within else None

        # 1) Milestone payment
        if (not is_discount) and (not is_penalty) and re.search(
            r"\b(milestone|đ[oợ]t|thanh\s*to[aá]n|payment|pay)\b", line, re.I
        ):
            pct = _parse_percent(line)
            conf = 0.4
            if pct is not None:
                conf += 0.3
            if due_date or within_days is not None:
                conf += 0.2
            conf = min(conf + 0.1, 1.0)
            out.append(
                {
                    "obligation_type": "milestone_payment",
                    "currency": currency or "VND",
                    "amount_value": None,
                    "amount_percent": pct,
                    "due_date": due_date,
                    "condition_text": line,
                    "confidence": conf,
                    "meta": {"within_days": within_days} if within_days is not None else None,
                }
            )

        # 2) Early payment discount
        if is_discount:
            pct = _parse_percent(line)
            conf = 0.4
            if pct is not None:
                conf += 0.3
            if due_date or within_days is not None:
                conf += 0.2
            if "early" in lower or "sớm" in lower or "som" in lower:
                conf += 0.1
            conf = min(conf, 1.0)
            out.append(
                {
                    "obligation_type": "early_payment_discount",
                    "currency": currency or "VND",
                    "amount_value": None,
                    "amount_percent": pct,
                    "due_date": due_date,
                    "condition_text": line,
                    "confidence": conf,
                    "meta": {"within_days": within_days} if within_days is not None else None,
                }
            )

        # 3) Late payment penalty
        if is_penalty:
            pct = _parse_percent(line)
            conf = 0.4
            if pct is not None:
                conf += 0.3
            if "per day" in lower or "/ngày" in lower or "/ngay" in lower:
                conf += 0.2
            if "late" in lower or "chậm" in lower or "cham" in lower:
                conf += 0.1
            conf = min(conf, 1.0)
            out.append(
                {
                    "obligation_type": "late_payment_penalty",
                    "currency": currency or "VND",
                    "amount_value": None,
                    "amount_percent": pct,
                    "due_date": due_date,
                    "condition_text": line,
                    "confidence": conf,
                    "meta": {"within_days": within_days} if within_days is not None else None,
                }
            )

    return out


def _normalize_trigger_key(text: str) -> str:
    t = " ".join(text.lower().split())
    # Strip volatile numbers so cross-source comparisons can detect conflicts.
    t = re.sub(r"[0-9]+", "#", t)
    t = re.sub(r"#+", "#", t)
    return t[:160]


def _evidence_strength(
    base_confidence: float,
    *,
    amount_present: bool,
    timepoint_present: bool,
    coords_present: bool,
    source_type: str,
) -> float:
    # Evidence-strength score, used for Tier gating (NOT ML accuracy).
    strength = float(base_confidence or 0.0)
    if amount_present:
        strength += 0.2
    if timepoint_present:
        strength += 0.2
    if coords_present:
        strength += 0.1

    # "Primary source" bias for product-mode (contract PDF over email).
    if source_type == "contract":
        strength *= float(getattr(settings, "obligation_primary_source_weight", 1.0))

    return max(0.0, min(1.0, strength))


def _missing_required_fields(
    *,
    amount_present: bool,
    timepoint_present: bool,
    trigger_present: bool,
    mode: str,
) -> list[str]:
    missing: list[str] = []
    if not amount_present:
        missing.append("amount")
    if mode == "strict" and not timepoint_present:
        # due_date or equivalent milestone timepoint
        missing.append("due_date")
    if not trigger_present:
        missing.append("trigger_condition")
    return missing


def _risk_level_max(a: str, b: str) -> str:
    order = {"low": 0, "medium": 1, "med": 1, "high": 2}
    aa = a if a in order else "medium"
    bb = b if b in order else "medium"
    return aa if order[aa] >= order[bb] else bb


def _classify_risk_level(
    obligation_type: str,
    *,
    has_email_source: bool,
    has_conflict: bool,
    missing_fields: list[str],
    ambiguous_condition: bool,
    currency: str,
) -> str:
    risk = "low"
    if obligation_type == "late_payment_penalty":
        risk = "high"
    elif obligation_type == "early_payment_discount":
        risk = "medium"

    if currency and currency.upper() != "VND":
        risk = _risk_level_max(risk, "high")
    if has_email_source:
        risk = _risk_level_max(risk, "medium")
    if ambiguous_condition:
        risk = _risk_level_max(risk, "high")
    if missing_fields:
        risk = _risk_level_max(risk, "medium")
    if has_conflict:
        risk = _risk_level_max(risk, "high")
    # normalize legacy "med"
    return "medium" if risk == "med" else risk


def _wf_contract_obligation(run_id: str) -> dict[str, Any]:
    payload = _run_payload(run_id)
    contract_files = list(payload.get("contract_files") or [])
    email_files = list(payload.get("email_files") or [])

    # Convenience for one-file demos (backward-ish compatible with attachment payloads).
    if not contract_files and payload.get("file_uri"):
        contract_files = [payload["file_uri"]]

    if not contract_files and not email_files:
        raise RuntimeError("payload.contract_files or payload.email_files is required")

    workdir = Path(tempfile.mkdtemp(prefix=f"agent-{run_id}-"))
    try:
        t_id = _task_start(
            run_id,
            "ingest_sources",
            {"contract_files": len(contract_files), "email_files": len(email_files)},
        )

        staged: list[dict[str, Any]] = []

        def _stage_uri(uri: str, source_type: str, idx: int) -> dict[str, Any]:
            if str(uri).startswith("s3://"):
                ref = parse_s3_uri(uri)
                suffix = Path(ref.key).suffix or ".bin"
                local_path = str(workdir / f"{source_type}-{idx}{suffix}")
                download_file(settings, ref, local_path)
            else:
                suffix = Path(str(uri)).suffix or ".bin"
                local_path = str(workdir / f"{source_type}-{idx}{suffix}")
                shutil.copyfile(uri, local_path)

            file_hash = sha256_file(local_path)
            content_type, _ = mimetypes.guess_type(local_path)
            return {
                "source_type": source_type,
                "source_uri": uri,
                "local_path": local_path,
                "ext": Path(local_path).suffix.lower(),
                "file_hash": file_hash,
                "size_bytes": int(Path(local_path).stat().st_size),
                "content_type": content_type,
                "stored_uri": uri if str(uri).startswith("s3://") else None,
            }

        for i, uri in enumerate(contract_files):
            staged.append(_stage_uri(str(uri), "contract", i))
        for i, uri in enumerate(email_files):
            staged.append(_stage_uri(str(uri), "email", i))

        contract_hashes = sorted([x["file_hash"] for x in staged if x["source_type"] == "contract"])
        email_hashes = sorted([x["file_hash"] for x in staged if x["source_type"] == "email"])
        case_key = payload.get("case_key") or make_idempotency_key("contract_case", contract_hashes, email_hashes)

        # Upload to drop bucket (stable key by hash) for local-file inputs.
        for item in staged:
            if item.get("stored_uri"):
                continue
            key = f"contract_cases/{case_key}/{item['source_type']}/{item['file_hash']}{item['ext']}"
            obj = upload_file(
                settings,
                settings.minio_bucket_drop,
                key,
                item["local_path"],
                content_type=item.get("content_type"),
            )
            item["stored_uri"] = obj.uri()

        _task_finish(run_id, "ingest_sources", "success", {"case_key": case_key, "sources": len(staged)})
        _db_log(run_id, t_id, "info", "contract_ingested", {"case_key": case_key, "sources": len(staged)})

        # Phase 1: upsert case + source rows
        with db_session(engine) as s:
            existing_case = s.execute(
                select(AgentContractCase).where(AgentContractCase.case_key == case_key)
            ).scalar_one_or_none()

            case_from_source_id: str | None = None
            for item in staged:
                src = s.execute(
                    select(AgentSourceFile).where(
                        (AgentSourceFile.file_hash == item["file_hash"])
                        & (AgentSourceFile.source_type == item["source_type"])
                    )
                ).scalar_one_or_none()
                if src and src.case_id:
                    case_from_source_id = src.case_id
                    break

            case = existing_case
            if not case and case_from_source_id:
                case = s.get(AgentContractCase, case_from_source_id)

            if not case:
                case = AgentContractCase(
                    case_id=new_uuid(),
                    case_key=case_key,
                    partner_name=payload.get("partner_name"),
                    partner_tax_id=payload.get("partner_tax_id"),
                    contract_code=payload.get("contract_code"),
                    status="open",
                    meta=payload.get("meta"),
                )
                s.add(case)
                s.flush()

            # Backfill from payload if present
            if payload.get("partner_name") and not case.partner_name:
                case.partner_name = payload["partner_name"]
            if payload.get("partner_tax_id") and not case.partner_tax_id:
                case.partner_tax_id = payload["partner_tax_id"]
            if payload.get("contract_code") and not case.contract_code:
                case.contract_code = payload["contract_code"]

            source_ids: list[str] = []
            for item in staged:
                src = s.execute(
                    select(AgentSourceFile).where(
                        (AgentSourceFile.file_hash == item["file_hash"])
                        & (AgentSourceFile.source_type == item["source_type"])
                    )
                ).scalar_one_or_none()
                if src:
                    if not src.case_id:
                        src.case_id = case.case_id
                    if not src.stored_uri and item.get("stored_uri"):
                        src.stored_uri = item["stored_uri"]
                    source_ids.append(src.source_id)
                    continue

                src = AgentSourceFile(
                    source_id=new_uuid(),
                    case_id=case.case_id,
                    source_type=item["source_type"],
                    source_uri=item["source_uri"],
                    stored_uri=item.get("stored_uri"),
                    file_hash=item["file_hash"],
                    size_bytes=item.get("size_bytes"),
                    content_type=item.get("content_type"),
                    meta=None,
                )
                s.add(src)
                s.flush()
                source_ids.append(src.source_id)

            case_id = case.case_id

        # Phase 2: extract contract text + parse emails (idempotent per source_id)
        contract_sources: list[dict[str, Any]] = []
        email_sources: list[dict[str, Any]] = []

        t_id = _task_start(run_id, "extract_contract_text", {"contracts": len(contract_files)})
        try:
            for item in staged:
                if item["source_type"] != "contract":
                    continue

                ext = Path(item["local_path"]).suffix.lower()
                engine_name = "unknown"
                pages_text: list[str] = []
                text = ""

                if ext == ".pdf":
                    engine_name = "pdfplumber"
                    try:
                        with pdfplumber.open(item["local_path"]) as pdf:
                            for page in pdf.pages:
                                pages_text.append(page.extract_text() or "")
                        text = "\n".join(t for t in pages_text if t.strip())
                    except Exception:
                        text = ""
                        pages_text = []

                    if not text.strip():
                        engine_name = "tesseract"
                        text = _ocr_pdf(item["local_path"], max_pages=settings.ocr_pdf_max_pages)
                        pages_text = [text]
                elif ext in {".png", ".jpg", ".jpeg", ".tif", ".tiff"}:
                    engine_name = "tesseract"
                    text = _ocr_image(item["local_path"])
                    pages_text = [text]
                else:
                    engine_name = "raw"
                    text = Path(item["local_path"]).read_text(encoding="utf-8", errors="ignore")
                    pages_text = [text]

                with db_session(engine) as s:
                    src = s.execute(
                        select(AgentSourceFile).where(
                            (AgentSourceFile.file_hash == item["file_hash"])
                            & (AgentSourceFile.source_type == item["source_type"])
                        )
                    ).scalar_one_or_none()
                    if not src:
                        continue

                    existing = s.execute(
                        select(AgentExtractedText).where(AgentExtractedText.source_id == src.source_id)
                    ).scalar_one_or_none()
                    if not existing:
                        s.add(
                            AgentExtractedText(
                                text_id=new_uuid(),
                                source_id=src.source_id,
                                engine=engine_name,
                                text=text,
                                page_confidence=None,
                            )
                        )
                    else:
                        # Keep the stored text stable for later reads.
                        text = existing.text

                    contract_sources.append(
                        {"source_id": src.source_id, "text": text, "pages_text": pages_text}
                    )
        finally:
            _task_finish(run_id, "extract_contract_text", "success", {"sources": len(contract_sources)})

        t_id = _task_start(run_id, "parse_emails", {"emails": len(email_files)})
        try:
            for item in staged:
                if item["source_type"] != "email":
                    continue

                subject, from_addr, to_addrs, clean_text = _parse_email_file(item["local_path"])

                with db_session(engine) as s:
                    src = s.execute(
                        select(AgentSourceFile).where(
                            (AgentSourceFile.file_hash == item["file_hash"])
                            & (AgentSourceFile.source_type == item["source_type"])
                        )
                    ).scalar_one_or_none()
                    if not src:
                        continue

                    existing = s.execute(
                        select(AgentEmailThread).where(AgentEmailThread.source_id == src.source_id)
                    ).scalar_one_or_none()
                    if not existing:
                        s.add(
                            AgentEmailThread(
                                thread_id=new_uuid(),
                                source_id=src.source_id,
                                subject=subject,
                                from_addr=from_addr,
                                to_addrs=to_addrs,
                                clean_text=clean_text or "",
                                highlights=None,
                            )
                        )
                    else:
                        # Keep the stored clean_text stable for later reads.
                        clean_text = existing.clean_text
                        subject = existing.subject

                    email_sources.append(
                        {
                            "source_id": src.source_id,
                            "subject": subject,
                            "from_addr": from_addr,
                            "to_addrs": to_addrs,
                            "clean_text": clean_text or "",
                        }
                    )
        finally:
            _task_finish(run_id, "parse_emails", "success", {"sources": len(email_sources)})

        combined_text = "\n".join([x["text"] for x in contract_sources] + [x["clean_text"] for x in email_sources])

        # Phase 3: extract obligations + evidence (3-tier gating uses evidence strength + conflicts)
        t_id = _task_start(run_id, "extract_obligations", {"text_len": len(combined_text)})
        groups: dict[str, dict[str, Any]] = {}
        obligations_created = 0
        evidence_created = 0
        group_signatures: list[str] = []
        try:
            for src in contract_sources:
                pages = src.get("pages_text") or [src.get("text") or ""]
                for page_no, page_text in enumerate(pages, start=1):
                    for line_no, raw_line in enumerate(str(page_text).splitlines(), start=1):
                        for cand in _extract_obligation_candidates(raw_line):
                            within_days = (cand.get("meta") or {}).get("within_days")
                            due_date = cand.get("due_date")
                            trigger_key = _normalize_trigger_key(str(cand.get("condition_text") or ""))
                            group_key = f"{cand['obligation_type']}:{trigger_key}"

                            amount_present = (cand.get("amount_value") is not None) or (
                                cand.get("amount_percent") is not None
                            )
                            timepoint_present = (due_date is not None) or (within_days is not None)
                            trigger_present = bool(str(cand.get("condition_text") or "").strip())
                            coords = {"page": page_no, "line": line_no}
                            strength = _evidence_strength(
                                float(cand.get("confidence") or 0.0),
                                amount_present=amount_present,
                                timepoint_present=timepoint_present,
                                coords_present=True,
                                source_type="contract",
                            )

                            groups.setdefault(
                                group_key,
                                {
                                    "obligation_type": cand["obligation_type"],
                                    "candidates": [],
                                },
                            )["candidates"].append(
                                {
                                    **cand,
                                    "source_id": src["source_id"],
                                    "source_type": "contract",
                                    "within_days": within_days,
                                    "coords": coords,
                                    "evidence_strength": strength,
                                    "amount_present": amount_present,
                                    "timepoint_present": timepoint_present,
                                    "trigger_present": trigger_present,
                                }
                            )

            for src in email_sources:
                subject = src.get("subject")
                for line_no, raw_line in enumerate(str(src.get("clean_text") or "").splitlines(), start=1):
                    for cand in _extract_obligation_candidates(raw_line):
                        within_days = (cand.get("meta") or {}).get("within_days")
                        due_date = cand.get("due_date")
                        trigger_key = _normalize_trigger_key(str(cand.get("condition_text") or ""))
                        group_key = f"{cand['obligation_type']}:{trigger_key}"

                        amount_present = (cand.get("amount_value") is not None) or (cand.get("amount_percent") is not None)
                        timepoint_present = (due_date is not None) or (within_days is not None)
                        trigger_present = bool(str(cand.get("condition_text") or "").strip())
                        coords = {"email_line": line_no, "subject": subject}
                        strength = _evidence_strength(
                            float(cand.get("confidence") or 0.0),
                            amount_present=amount_present,
                            timepoint_present=timepoint_present,
                            coords_present=True,
                            source_type="email",
                        )

                        groups.setdefault(
                            group_key,
                            {
                                "obligation_type": cand["obligation_type"],
                                "candidates": [],
                            },
                        )["candidates"].append(
                            {
                                **cand,
                                "source_id": src["source_id"],
                                "source_type": "email",
                                "within_days": within_days,
                                "coords": coords,
                                "evidence_strength": strength,
                                "amount_present": amount_present,
                                "timepoint_present": timepoint_present,
                                "trigger_present": trigger_present,
                            }
                        )

            mode = str(getattr(settings, "obligation_required_fields", "strict") or "strict")
            for group_key, g in groups.items():
                candidates = list(g.get("candidates") or [])
                if not candidates:
                    continue

                def _field_conflicts(
                    cands: list[dict[str, Any]], field: str
                ) -> list[dict[str, Any]] | None:
                    vals: dict[str, list[dict[str, Any]]] = {}
                    for c in cands:
                        v = c.get(field)
                        if v is None:
                            continue
                        key = str(v)
                        vals.setdefault(key, []).append(
                            {"source_id": c["source_id"], "source_type": c["source_type"]}
                        )
                    if len(vals) <= 1:
                        return None
                    return [{"value": k, "sources": srcs} for k, srcs in vals.items()]

                conflicts: dict[str, Any] = {}
                for f in ["amount_value", "amount_percent", "due_date", "within_days"]:
                    c = _field_conflicts(candidates, f)
                    if c:
                        conflicts[f] = c

                has_conflict = bool(conflicts)
                has_email_source = any(c.get("source_type") == "email" for c in candidates)

                # Choose a "best" candidate for display (prefer contract, then strength).
                candidates_sorted = sorted(
                    candidates,
                    key=lambda c: (
                        1 if c.get("source_type") == "contract" else 0,
                        float(c.get("evidence_strength") or 0.0),
                    ),
                    reverse=True,
                )
                best = candidates_sorted[0]
                currency = (best.get("currency") or "VND").upper()

                # If a field conflicts, keep it unset and surface the conflict in meta.
                amount_value = None if "amount_value" in conflicts else best.get("amount_value")
                amount_percent = None if "amount_percent" in conflicts else best.get("amount_percent")
                due_date = None if "due_date" in conflicts else best.get("due_date")
                within_days = None if "within_days" in conflicts else best.get("within_days")

                amount_present = (amount_value is not None) or (amount_percent is not None)
                timepoint_present = (due_date is not None) or (within_days is not None)
                trigger_present = bool(str(best.get("condition_text") or "").strip())

                missing_fields = _missing_required_fields(
                    amount_present=amount_present,
                    timepoint_present=timepoint_present,
                    trigger_present=trigger_present,
                    mode=mode,
                )

                strong_evidence = any(
                    bool(c.get("amount_present"))
                    and bool(c.get("timepoint_present"))
                    and bool(c.get("trigger_present"))
                    and bool(c.get("coords"))
                    for c in candidates
                )

                cond_lower = str(best.get("condition_text") or "").lower()
                ambiguous_condition = any(
                    k in cond_lower
                    for k in [
                        "tbd",
                        "to be discussed",
                        "subject to",
                        "pending",
                        "nghiem thu",
                        "nghiệm thu",
                        "acceptance",
                        "tranh chấp",
                        "dispute",
                    ]
                )

                risk_level = _classify_risk_level(
                    best.get("obligation_type") or g.get("obligation_type") or "unknown",
                    has_email_source=has_email_source,
                    has_conflict=has_conflict,
                    missing_fields=missing_fields,
                    ambiguous_condition=ambiguous_condition,
                    currency=currency,
                )

                signature = sha256_text(json_dumps_canonical([case_key, group_key]))
                meta: dict[str, Any] = {
                    "group_key": group_key,
                    "within_days": within_days,
                    "missing_fields": missing_fields,
                    "strong_evidence": bool(strong_evidence),
                    "conflicts": conflicts or None,
                    "sources": [
                        {"source_id": c["source_id"], "source_type": c["source_type"]} for c in candidates_sorted
                    ],
                }

                max_strength = max(float(c.get("evidence_strength") or 0.0) for c in candidates_sorted)

                with db_session(engine) as s:
                    ob = s.execute(
                        select(AgentObligation).where(AgentObligation.signature == signature)
                    ).scalar_one_or_none()
                    if not ob:
                        ob = AgentObligation(
                            obligation_id=new_uuid(),
                            case_id=case_id,
                            obligation_type=best.get("obligation_type") or g.get("obligation_type") or "unknown",
                            currency=currency,
                            amount_value=amount_value,
                            amount_percent=amount_percent,
                            due_date=due_date,
                            condition_text=str(best.get("condition_text") or "")[:4000],
                            confidence=max_strength,
                            risk_level=risk_level,
                            signature=signature,
                            meta=meta,
                        )
                        s.add(ob)
                        s.flush()
                        obligations_created += 1
                    else:
                        ob.confidence = max(float(ob.confidence or 0.0), max_strength)
                        ob.risk_level = risk_level
                        ob.meta = meta
                        # only fill missing scalar fields if we have a non-conflicting value
                        if amount_value is not None and ob.amount_value is None:
                            ob.amount_value = amount_value
                        if amount_percent is not None and ob.amount_percent is None:
                            ob.amount_percent = amount_percent
                        if due_date is not None and ob.due_date is None:
                            ob.due_date = due_date
                        if ob.condition_text and ob.condition_text.strip():
                            pass
                        else:
                            ob.condition_text = str(best.get("condition_text") or "")[:4000]

                    for ev in candidates_sorted:
                        evidence_type = "email" if ev.get("source_type") == "email" else "quote"
                        snippet = str(ev.get("condition_text") or "")[:2000]
                        evidence_id = make_idempotency_key(
                            "obligation_evidence",
                            signature,
                            ev["source_id"],
                            evidence_type,
                            snippet,
                            ev.get("coords"),
                        )[:36]
                        if s.get(AgentObligationEvidence, evidence_id):
                            continue
                        s.add(
                            AgentObligationEvidence(
                                evidence_id=evidence_id,
                                obligation_id=ob.obligation_id,
                                source_id=ev["source_id"],
                                evidence_type=evidence_type,
                                snippet=snippet,
                                meta={
                                    "coords": ev.get("coords"),
                                    "source_type": ev.get("source_type"),
                                    "evidence_strength": float(ev.get("evidence_strength") or 0.0),
                                    "group_key": group_key,
                                },
                            )
                        )
                        evidence_created += 1

                group_signatures.append(signature)
        finally:
            _task_finish(
                run_id,
                "extract_obligations",
                "success",
                {"obligations": obligations_created, "evidence": evidence_created},
            )

        # Phase 4: reconcile ERPX (read-only) — persist agent_erpx_links
        t_id = _task_start(run_id, "reconcile_erpx")
        erpx_links_created = 0
        try:
            contract_meta = _extract_contract_meta(combined_text)
            with db_session(engine) as s:
                case = s.get(AgentContractCase, case_id)
                if case:
                    if contract_meta.get("partner_tax_id") and not case.partner_tax_id:
                        case.partner_tax_id = contract_meta["partner_tax_id"]
                    if contract_meta.get("contract_code") and not case.contract_code:
                        case.contract_code = contract_meta["contract_code"]

            client = ErpXClient(settings)
            try:
                erpx_contracts = client.get_contracts(
                    partner_id=contract_meta.get("partner_tax_id"),
                )
                erpx_payments = client.get_payments(
                    contract_id=erpx_contracts[0]["contract_id"] if erpx_contracts else None,
                )
            except Exception as e:
                _db_log(run_id, t_id, "warn", "erpx_read_partial", {"error": str(e)})
                erpx_contracts = []
                erpx_payments = []
            finally:
                client.close()

            # Link ERPX contracts + payments to this case (idempotent by signature)
            with db_session(engine) as s:
                obs = (
                    s.execute(
                        select(AgentObligation).where(AgentObligation.case_id == case_id)
                    )
                    .scalars()
                    .all()
                )

                for ec in erpx_contracts:
                    existing = s.execute(
                        select(AgentErpXLink).where(
                            (AgentErpXLink.case_id == case_id)
                            & (AgentErpXLink.erpx_object_type == "contract")
                            & (AgentErpXLink.erpx_object_id == str(ec.get("contract_id", "")))
                        )
                    ).scalar_one_or_none()
                    if not existing:
                        s.add(
                            AgentErpXLink(
                                link_id=new_uuid(),
                                case_id=case_id,
                                obligation_id=None,
                                erpx_object_type="contract",
                                erpx_object_id=str(ec.get("contract_id", "")),
                                match_confidence=0.8,
                                meta={"partner_name": ec.get("partner_name"), "contract_code": ec.get("contract_code")},
                            )
                        )
                        erpx_links_created += 1

                for ep in erpx_payments:
                    existing = s.execute(
                        select(AgentErpXLink).where(
                            (AgentErpXLink.case_id == case_id)
                            & (AgentErpXLink.erpx_object_type == "payment")
                            & (AgentErpXLink.erpx_object_id == str(ep.get("payment_id", "")))
                        )
                    ).scalar_one_or_none()
                    if not existing:
                        # Try to match payment to an obligation by due date or amount
                        matched_ob_id = None
                        best_conf = 0.5
                        for ob in obs:
                            if ob.amount_value and ep.get("amount"):
                                ratio = min(float(ob.amount_value), float(ep["amount"])) / max(
                                    float(ob.amount_value), float(ep["amount"]), 1.0
                                )
                                if ratio > 0.95:
                                    matched_ob_id = ob.obligation_id
                                    best_conf = ratio
                                    break
                        s.add(
                            AgentErpXLink(
                                link_id=new_uuid(),
                                case_id=case_id,
                                obligation_id=matched_ob_id,
                                erpx_object_type="payment",
                                erpx_object_id=str(ep.get("payment_id", "")),
                                match_confidence=best_conf,
                                meta={"amount": ep.get("amount"), "date": ep.get("date")},
                            )
                        )
                        erpx_links_created += 1

            _db_log(run_id, t_id, "info", "erpx_reconciled", {
                "contracts": len(erpx_contracts),
                "payments": len(erpx_payments),
                "links_created": erpx_links_created,
            })
        finally:
            _task_finish(run_id, "reconcile_erpx", "success", {"erpx_links_created": erpx_links_created})

        # Phase 5: create proposals (3-tier gating + maker-checker metadata)
        t_id = _task_start(run_id, "create_proposals")
        try:
            proposals_created = 0
            tier_counts = {1: 0, 2: 0, 3: 0}

            with db_session(engine) as s:
                run = s.get(AgentRun, run_id)
                created_by = (getattr(run, "requested_by", None) or "").strip() or "system"

                threshold = float(getattr(settings, "obligation_confidence_threshold", 0.8))
                mode = str(getattr(settings, "obligation_required_fields", "strict") or "strict")
                conflict_policy = str(getattr(settings, "obligation_conflict_policy", "drop_to_tier2"))

                obs = (
                    s.execute(
                        select(AgentObligation)
                        .where(AgentObligation.case_id == case_id)
                        .order_by(AgentObligation.created_at.desc())
                    )
                    .scalars()
                    .all()
                )

                # Evidence pack: zip (index.json + extracted texts), idempotent by (issue_key, version=1)
                issue_key = f"contract_obligation:{case_id}"
                ev_pack = s.execute(
                    select(AgentEvidencePack).where(
                        (AgentEvidencePack.issue_key == issue_key) & (AgentEvidencePack.version == 1)
                    )
                ).scalar_one_or_none()
                pack_uri = ev_pack.pack_uri if ev_pack else None
                if not ev_pack:
                    index_json = {
                        "case_id": case_id,
                        "case_key": case_key,
                        "generated_at": utcnow().isoformat(),
                        "obligations": [
                            {
                                "obligation_id": ob.obligation_id,
                                "signature": ob.signature,
                                "obligation_type": ob.obligation_type,
                                "currency": ob.currency,
                                "amount_value": ob.amount_value,
                                "amount_percent": ob.amount_percent,
                                "due_date": str(ob.due_date) if ob.due_date else None,
                                "condition_text": ob.condition_text,
                                "confidence": float(ob.confidence or 0.0),
                                "risk_level": getattr(ob, "risk_level", None),
                                "meta": ob.meta,
                            }
                            for ob in obs
                        ],
                    }

                    pack_path = str(workdir / "contract_obligation_evidence_pack.zip")
                    with zipfile.ZipFile(pack_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
                        z.writestr("index.json", json_dumps_canonical(index_json))
                        for src in contract_sources:
                            z.writestr(f"sources/{src['source_id']}.txt", str(src.get("text") or ""))
                        for src in email_sources:
                            z.writestr(
                                f"sources/{src['source_id']}.eml.txt", str(src.get("clean_text") or "")
                            )

                    key = f"evidence/contract_obligation/{case_key}/pack_v1.zip"
                    obj = upload_file(
                        settings,
                        settings.minio_bucket_evidence,
                        key,
                        pack_path,
                        content_type="application/zip",
                    )
                    pack_uri = obj.uri()
                    ev_pack = AgentEvidencePack(
                        id=new_uuid(),
                        issue_key=issue_key,
                        version=1,
                        pack_uri=pack_uri,
                        index_json=index_json,
                        run_id=run_id,
                    )
                    s.add(ev_pack)

                def _tier_for_obligation(ob: AgentObligation) -> tuple[int, list[str], dict | None]:
                    meta = ob.meta if isinstance(ob.meta, dict) else {}
                    within_days = meta.get("within_days")
                    strong_evidence = bool(meta.get("strong_evidence"))
                    conflicts = meta.get("conflicts")

                    amount_present = (ob.amount_value is not None) or (ob.amount_percent is not None)
                    timepoint_present = (ob.due_date is not None) or (within_days is not None)
                    trigger_present = bool((ob.condition_text or "").strip())
                    missing_fields = _missing_required_fields(
                        amount_present=amount_present,
                        timepoint_present=timepoint_present,
                        trigger_present=trigger_present,
                        mode=mode,
                    )

                    if conflicts and conflict_policy == "drop_to_tier2":
                        return 2, missing_fields, conflicts
                    if float(ob.confidence or 0.0) < threshold:
                        return 3, missing_fields, conflicts
                    if missing_fields:
                        return 2, missing_fields, conflicts
                    if not strong_evidence:
                        return 2, missing_fields, conflicts
                    return 1, missing_fields, conflicts

                for ob in obs:
                    tier, missing_fields, conflicts = _tier_for_obligation(ob)
                    tier_counts[tier] = tier_counts.get(tier, 0) + 1

                    risk_level = (getattr(ob, "risk_level", None) or "medium").strip().lower()
                    if risk_level == "med":
                        risk_level = "medium"

                    base_details = {
                        "tier": tier,
                        "risk_level": risk_level,
                        "evidence_pack_uri": pack_uri,
                        "obligation_signature": ob.signature,
                        "missing_fields": missing_fields,
                        "conflicts": conflicts,
                        "note": "Draft outside ERPX core. Agent does not post entries to ERPX.",
                    }

                    if tier == 1:
                        # Tier 1: draft proposals + evidence pack
                        reminder_key = make_idempotency_key(
                            "contract_proposal", ob.signature, "reminder", tier
                        )
                        existing = s.execute(
                            select(AgentProposal).where(AgentProposal.proposal_key == reminder_key)
                        ).scalar_one_or_none()
                        if not existing:
                            details = {
                                **base_details,
                                "due_date": str(ob.due_date) if ob.due_date else None,
                                "within_days": (ob.meta or {}).get("within_days") if isinstance(ob.meta, dict) else None,
                            }
                            p = AgentProposal(
                                proposal_id=new_uuid(),
                                case_id=case_id,
                                obligation_id=ob.obligation_id,
                                proposal_type="reminder",
                                title=f"Reminder (Tier1): {ob.obligation_type}",
                                summary=ob.condition_text[:2000],
                                details=details,
                                risk_level=risk_level,
                                confidence=float(ob.confidence or 0.0),
                                status="draft",
                                created_by=created_by,
                                tier=1,
                                evidence_summary_hash=sha256_text(json_dumps_canonical(details)),
                                proposal_key=reminder_key,
                                run_id=run_id,
                            )
                            s.add(p)
                            s.flush()
                            s.add(
                                AgentAuditLog(
                                    audit_id=new_uuid(),
                                    actor_user_id=created_by,
                                    action="proposal.create",
                                    object_type="proposal",
                                    object_id=p.proposal_id,
                                    before=None,
                                    after={
                                        "proposal_id": p.proposal_id,
                                        "case_id": p.case_id,
                                        "obligation_id": p.obligation_id,
                                        "proposal_type": p.proposal_type,
                                        "proposal_key": p.proposal_key,
                                        "tier": 1,
                                        "risk_level": risk_level,
                                    },
                                    run_id=run_id,
                                )
                            )
                            proposals_created += 1

                        if ob.obligation_type != "milestone_payment":
                            continue

                        accrual_key = make_idempotency_key(
                            "contract_proposal", ob.signature, "accrual_template", tier
                        )
                        existing = s.execute(
                            select(AgentProposal).where(AgentProposal.proposal_key == accrual_key)
                        ).scalar_one_or_none()
                        if existing:
                            continue

                        details = {
                            **base_details,
                            "template": {
                                "currency": ob.currency,
                                "amount_percent": ob.amount_percent,
                                "amount_value": ob.amount_value,
                                "due_date": str(ob.due_date) if ob.due_date else None,
                            }
                        }
                        p = AgentProposal(
                            proposal_id=new_uuid(),
                            case_id=case_id,
                            obligation_id=ob.obligation_id,
                            proposal_type="accrual_template",
                            title="Accrual/deferral template (Tier1, review & approve)",
                            summary=ob.condition_text[:2000],
                            details=details,
                            risk_level=risk_level,
                            confidence=float(ob.confidence or 0.0),
                            status="draft",
                            created_by=created_by,
                            tier=1,
                            evidence_summary_hash=sha256_text(json_dumps_canonical(details)),
                            proposal_key=accrual_key,
                            run_id=run_id,
                        )
                        s.add(p)
                        s.flush()
                        s.add(
                            AgentAuditLog(
                                audit_id=new_uuid(),
                                actor_user_id=created_by,
                                action="proposal.create",
                                object_type="proposal",
                                object_id=p.proposal_id,
                                before=None,
                                after={
                                    "proposal_id": p.proposal_id,
                                    "case_id": p.case_id,
                                    "obligation_id": p.obligation_id,
                                    "proposal_type": p.proposal_type,
                                    "proposal_key": p.proposal_key,
                                    "tier": 1,
                                    "risk_level": risk_level,
                                },
                                run_id=run_id,
                            )
                        )
                        proposals_created += 1
                        continue

                    if tier == 2:
                        # Tier 2: summary + quick confirm (no accounting template)
                        p_key = make_idempotency_key(
                            "contract_proposal", ob.signature, "review_confirm", tier
                        )
                        existing = s.execute(
                            select(AgentProposal).where(AgentProposal.proposal_key == p_key)
                        ).scalar_one_or_none()
                        if existing:
                            continue

                        details = {
                            **base_details,
                            "action_required": "confirm_fields",
                            "hint": "Resolve missing/uncertain fields or conflicts, then re-run.",
                        }
                        p = AgentProposal(
                            proposal_id=new_uuid(),
                            case_id=case_id,
                            obligation_id=ob.obligation_id,
                            proposal_type="review_confirm",
                            title=f"Confirm obligation (Tier2): {ob.obligation_type}",
                            summary=ob.condition_text[:2000],
                            details=details,
                            risk_level=risk_level,
                            confidence=float(ob.confidence or 0.0),
                            status="pending",
                            created_by=created_by,
                            tier=2,
                            evidence_summary_hash=sha256_text(json_dumps_canonical(details)),
                            proposal_key=p_key,
                            run_id=run_id,
                        )
                        s.add(p)
                        s.flush()
                        s.add(
                            AgentAuditLog(
                                audit_id=new_uuid(),
                                actor_user_id=created_by,
                                action="proposal.create",
                                object_type="proposal",
                                object_id=p.proposal_id,
                                before=None,
                                after={
                                    "proposal_id": p.proposal_id,
                                    "case_id": p.case_id,
                                    "obligation_id": p.obligation_id,
                                    "proposal_type": p.proposal_type,
                                    "proposal_key": p.proposal_key,
                                    "tier": 2,
                                    "risk_level": risk_level,
                                },
                                run_id=run_id,
                            )
                        )
                        proposals_created += 1
                        continue

                    # Tier 3: missing data (no inference)
                    p_key = make_idempotency_key(
                        "contract_proposal", ob.signature, "missing_data", tier
                    )
                    existing = s.execute(
                        select(AgentProposal).where(AgentProposal.proposal_key == p_key)
                    ).scalar_one_or_none()
                    if existing:
                        continue

                    details = {
                        **base_details,
                        "action_required": "provide_missing_data",
                        "hint": "Provide missing amount/date/trigger evidence (PDF/email excerpt) and re-run.",
                    }
                    p = AgentProposal(
                        proposal_id=new_uuid(),
                        case_id=case_id,
                        obligation_id=ob.obligation_id,
                        proposal_type="missing_data",
                        title=f"Missing data (Tier3): {ob.obligation_type}",
                        summary=ob.condition_text[:2000],
                        details=details,
                        risk_level=risk_level,
                        confidence=float(ob.confidence or 0.0),
                        status="pending",
                        created_by=created_by,
                        tier=3,
                        evidence_summary_hash=sha256_text(json_dumps_canonical(details)),
                        proposal_key=p_key,
                        run_id=run_id,
                    )
                    s.add(p)
                    s.flush()
                    s.add(
                        AgentAuditLog(
                            audit_id=new_uuid(),
                            actor_user_id=created_by,
                            action="proposal.create",
                            object_type="proposal",
                            object_id=p.proposal_id,
                            before=None,
                            after={
                                "proposal_id": p.proposal_id,
                                "case_id": p.case_id,
                                "obligation_id": p.obligation_id,
                                "proposal_type": p.proposal_type,
                                "proposal_key": p.proposal_key,
                                "tier": 3,
                                "risk_level": risk_level,
                            },
                            run_id=run_id,
                        )
                    )
                    proposals_created += 1

                if not obs:
                    # No obligations extracted at all -> Tier3 case-level output.
                    p_key = make_idempotency_key("contract_case", case_key, "missing_data")
                    existing = s.execute(
                        select(AgentProposal).where(AgentProposal.proposal_key == p_key)
                    ).scalar_one_or_none()
                    if not existing:
                        details = {
                            "tier": 3,
                            "risk_level": "medium",
                            "evidence_pack_uri": pack_uri,
                            "missing_fields": ["amount", "due_date", "trigger_condition"],
                            "action_required": "provide_source_documents",
                        }
                        p = AgentProposal(
                            proposal_id=new_uuid(),
                            case_id=case_id,
                            obligation_id=None,
                            proposal_type="missing_data",
                            title="Missing data (Tier3): no obligations extracted",
                            summary="No obligation clause was extracted with sufficient confidence.",
                            details=details,
                            risk_level="medium",
                            confidence=0.0,
                            status="pending",
                            created_by=created_by,
                            tier=3,
                            evidence_summary_hash=sha256_text(json_dumps_canonical(details)),
                            proposal_key=p_key,
                            run_id=run_id,
                        )
                        s.add(p)
                        proposals_created += 1

            _update_run(
                run_id,
                cursor_out={
                    "case_id": case_id,
                    "case_key": case_key,
                    "obligations": len(obs),
                    "proposals_created": proposals_created,
                    "tiers": tier_counts,
                },
            )
        finally:
            _task_finish(run_id, "create_proposals", "success")

        return {
            "case_id": case_id,
            "case_key": case_key,
            "sources": len(staged),
            "obligations": len(group_signatures),
        }
    finally:
        shutil.rmtree(workdir, ignore_errors=True)
